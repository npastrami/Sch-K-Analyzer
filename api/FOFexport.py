import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
import numpy as np
import pandas as pd
import re
from io import BytesIO
from itemcodeoffsets import keyword_to_offset_dict

def process_FOF(workbook, fof_sheets):   
        mappings = {
            "partnership ein": 7,
            "partnership name": 8,
            "partnership address": 9,
            "partnership city": 10,
            "partnership state": 11,
            "partnership zip code": 12,
            "partner's ssn/tin": 17,
            "partner name": 18,
            "partner address": 19,
            "partner city": 20,
            "partner state": 21,
            "partner zip": 22,
            "general partner or llc member-manager": 23,
            "limited partner or other llc member": 24,
            "domestic partner": 25,
            "foreign partner": 26,
            "if the partner is a disregarded entity": 27,
            "de partner tin": 28,
            "de partner name": 29,
            "de partner type": 30,
            "partner share of profit beginning": 31,
            "partner share of profit ending": 32,
            "partner share of loss beginning": 33,
            "partner share of loss ending": 34,
            "partner share of capital beginning": 35,
            "partner share of capital ending": 36,
            "nonrecourse beginning": 38,
            "nonrecourse ending": 39,
            "qualified nonrecourse financing beginning": 40,
            "qualified nonrecourse financing ending": 41,
            "recourse beginning": 42,
            "recourse ending": 43,
            "beginning capital account": 45,
            "capital contributed during the year": 46,
            "current year net income": 47,
            "other increases": 48,
            "withdrawls and distributions": 49,
            "ending capital account": 50,
            "partner's share of net unrecognized section": 52,
            "ordinary business income": 55,
            "net rental real estate income": 56, 
            "other net rental income": 57,
            "guarenteed payments for services": 58,
            "guarenteed payments for capital": 59,
            "total guarenteed payments": 60, 
            "interest income": 61,                       
            "ordinary dividends": 63, 
            "qualified dividends": 64, 
            "dividend equivalents": 65, 
            "royalties": 66,                       
            "net short-term capital gain": 67, 
            "net long-term capital gain": 68,                      
            "collectibles": 69, 
            "unrecaptured section": 70,                        
            "net section": 71, 
            "other income": 71, 
            "section": 97,                         
            "other deductions": 97, 
            "self-employment earnings": 135, 
            "credits": 139, 
            "alternatice minimum tax": 199, 
            "tax-exempt income and nondeductible expenses": 205, 
            "distributions": 208, 
            "other information": 211, 
            "foreign taxes paid or accrued": 287,  
        }
        
        target_worksheet = workbook["Sheet1"]
        starting_col = 5

        for sheet_index, fof_sheet in enumerate(fof_sheets, start=starting_col):
            # Extracting the sheet name and removing the 'FOF_' prefix
            sheet_title = fof_sheet.title.replace("FOF_", "")
            col_letter = get_column_letter(sheet_index)
            target_worksheet[f'{col_letter}7'] = sheet_title
            
            #list of lists to store items with footnotes
            items_with_footnotes = []
            items_with_target_row = []

            for row in fof_sheet.iter_rows(min_row=1):
                keyword_cell = row[0].value
                if keyword_cell and any(keyword in keyword_cell for keyword in mappings.keys()):
                    # match base keyword
                    base_keyword = re.match(r'([^\d]+)\s*(\d*)', keyword_cell).groups()[0]
                    base_keyword = base_keyword.strip()
                    # fetch item code associated with keyword instance
                    item_code = row[1].value

                    if item_code and '*' in item_code:
                        item_code = item_code.replace('*', '')
                        items_with_footnotes.append([base_keyword, item_code, int(sheet_index)])
                    if base_keyword in keyword_to_offset_dict:

                        offset_dict = keyword_to_offset_dict[base_keyword]

                        if item_code in offset_dict:
                            offset = offset_dict[item_code]
                            target_row = int(mappings[base_keyword]) + offset + 1  # Convert to int
                            print(f'Target row for {base_keyword} with item code {item_code} is {target_row}')
                        else:
                            invalid_item_codes = { base_keyword: item_code }
                            print(f'This Keyword-Item Code pairing is invalid: {invalid_item_codes}')
                            continue
                    else:
                        # For keywords without offsets
                        target_row = int(mappings[base_keyword]) + 1  # Convert to int
                    print(f'final target row: {target_row},for {base_keyword} with item code {item_code} ')
                    # append target row to items_with_target_row matched with base keyword and item code
                    for item in items_with_footnotes:
                        if item[0] == base_keyword and item[1] == item_code:
                            items_with_target_row.append(item + [target_row])  # Create a new list with target_row
                    
                    
                    amount_cell = row[2].value
                    target_worksheet.cell(row=target_row, column=sheet_index, value=amount_cell)  
            print(f'all items with footnotes: {items_with_footnotes}')
            
            for base_keyword, item_code, sheet_index, target_row in items_with_target_row:
                # Ensure target_row and sheet_index are integers before calling the cell method
                target_row = int(target_row) if isinstance(target_row, str) else target_row
                sheet_index = int(sheet_index) if isinstance(sheet_index, str) else sheet_index
                cell_to_highlight = target_worksheet.cell(row=target_row, column=sheet_index)
                cell_to_highlight.fill = openpyxl.styles.PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                        
        for col in range(1, target_worksheet.max_column + 1):
                col_letter = get_column_letter(col)
                target_worksheet.column_dimensions[col_letter].width = 25  # Approximate conversion
                
        # # Make Totals Column for Columns E-Z in Column D
        # for row in target_worksheet.iter_rows(min_row=57, max_row=290, min_col=4, max_col=4):
        #     for cell in row:
        #         cell.value = f'=SUM({get_column_letter(cell.column)}7:{get_column_letter(cell.column)}56)'
        
        for row_num in range(56, 291):  # End range is exclusive, so 291 to include row 290
            sum_value = 0  # Initialize sum for the current row
            for col_num in range(5, target_worksheet.max_column + 1):  # Start from column E, which is the 5th column
                cell = target_worksheet.cell(row=row_num, column=col_num)
                cell_value = cell.value
                # Check if the cell contains a string that can be converted to an integer
                if isinstance(cell_value, str):
                    # Attempt to convert string to integer if it's not empty and is numeric
                    try:
                        if cell_value.strip().isdigit():  # Check if the string is numeric
                            cell_value = int(cell_value)
                        else:
                            continue
                    except ValueError:
                        continue
                if isinstance(cell_value, (int, float)):
                    sum_value += cell_value # if sum_value = 0 then make the cell value empty
        
            # # lets highlight the target cells of base keywords with item footnotes
            # for row in range(55, 291):
            #     for col in range(starting_col, target_worksheet.max_column + 1):
            #         cell = target_worksheet.cell(row=row, column=col)
            #         for keyword, item_code in items_with_footnotes:
            #             if cell.value is not None and keyword in items_with_footnotes:
            #                 cell.fill = openpyxl.styles.PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                    
            # Update Column D with the calculated sum for the current row, converting sum to string, make 0 totals blank
            if sum_value != 0:
                target_worksheet.cell(row=row_num, column=4, value=str(sum_value))  # Column D is the 4th column
            else:
                target_worksheet.cell(row=row_num, column=4, value="")
        
        # Set text wrap for row 7
        for col in range(1, target_worksheet.max_column + 1):
            cell = target_worksheet.cell(row=7, column=col)
            cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

        # Rename the worksheet
        target_worksheet.title = "FOF_Data"
    