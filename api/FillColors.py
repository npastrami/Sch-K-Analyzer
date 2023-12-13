import re
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def fill_cells_with_colors(sheet, conf_sheet):
    # Define your color ranges
    color_ranges = [
        (0.0, 0.1, 'FFFF0000'),  # Red
        (0.1, 0.2, 'FFFF4C4C'),
        (0.2, 0.3, 'FFFF9999'),
        (0.3, 0.4, 'FFFFCC99'),
        (0.4, 0.5, 'FFFFFF99'),
        (0.5, 0.6, 'FFCCFF99'),
        (0.6, 0.7, 'FF99FF99'),
        (0.7, 0.8, 'FF66FF99'),
        (0.8, 0.9, 'FF33FF99'),
        (0.9, 1.0, 'FF00FF00'),  # Green
    ]

    # Define the font, alignment, and border style
    font = Font(name='Calibri', bold=True, italic=True, size=11, color="FF0000CC")
    alignment = Alignment(horizontal="center", vertical="center")
    border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

    fill_color_other = 'FFD3D3D3'  # Add the fill color for other cells here
    fill = PatternFill(fill_type="solid", fgColor=fill_color_other)

    # Define cell size (height and width)
    cell_height = 20  # replace with your value
    cell_width = 25  # replace with your value

    for col in range(5, sheet.max_column + 1):
        for row in range(22, 145):
            conf_cell = conf_sheet.cell(row=row, column=col)
            if conf_cell.value is not None:  # Check if the cell has a confidence score
                conf_score = float(conf_cell.value)
                for low, high, color in color_ranges:
                    if low <= conf_score < high:
                        fill_color = color
                        break
            else:
            # Check if the corresponding cell in 'sheet' has a value with alphabetical or numerical characters
                cell_value = sheet.cell(row=row, column=col).value
                if cell_value is not None and re.search(r'[A-Za-z0-9]', cell_value):
                    fill_color = 'FFADD8E6'  # Light blue for cells with a value but no confidence score
                else:
                    continue

            # Define the cell fill
            fill_conf = PatternFill(fill_type="solid", fgColor=fill_color)

            # Apply the fill to the cell in 'sheet'
            cell = sheet.cell(row=row, column=col)
            cell.fill = fill_conf

            # Apply cell size
            sheet.row_dimensions[row].height = cell_height
            sheet.column_dimensions[get_column_letter(col)].width = cell_width

    # Apply the formatting to the first 4 columns and rows 1 to 21
    for col in range(1, 5):
        for row in range(7, 22):
            cell = sheet.cell(row=row, column=col)
            cell.font = font
            cell.alignment = alignment
            cell.border = border

            # Apply cell size
            sheet.row_dimensions[row].height = cell_height
            sheet.column_dimensions[get_column_letter(col)].width = cell_width

    # Apply the formatting to columns 5 to max_column and rows 1 to 21
    for col in range(5, sheet.max_column + 1):
        for row in range(1, 22):
            cell = sheet.cell(row=row, column=col)
            cell.alignment = alignment
            cell.border = border

            # Apply cell size
            sheet.row_dimensions[row].height = cell_height
            sheet.column_dimensions[get_column_letter(col)].width = cell_width

    # Apply the formatting to the first 4 columns and rows 145 to max_row
    for col in range(1, 5):
        for row in range(145, sheet.max_row + 1):
            cell = sheet.cell(row=row, column=col)
            cell.alignment = alignment
            cell.border = border
            cell.fill = fill

            # Apply cell size
            sheet.row_dimensions[row].height = cell_height
            sheet.column_dimensions[get_column_letter(col)].width = cell_width